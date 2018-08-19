import click

from sqlalchemy_utils import database_exists, create_database
from openpyxl import load_workbook, Workbook

from ami.app import create_app
from ami.extensions import db
from ami.blueprints.user.models import User
from ami.blueprints.imei.models import Imei, Sim
from ami.blueprints.meter.models import Meter
from ami.blueprints.data.models import AlertData
from datetime import datetime
import os
import random

from faker import Faker

from lib.util_datetime import tzware_datetime

# Create an app context for the database connection.
app = create_app()
db.app = app

fake = Faker()


def _log_status(count, model_label):
    """
    Log the output of how many records were created.

    :param count: Amount created
    :type count: int
    :param model_label: Name of the model
    :type model_label: str
    :return: None
    """
    click.echo('Created {0} {1}'.format(count, model_label))

    return None


@click.group()
def cli():
    """ Run PostgreSQL related tasks. """
    pass


@click.command()
@click.option('--with-testdb/--no-with-testdb', default=False,
              help='Create a test db too?')
def init(with_testdb):
    """
    Initialize the database.

    :param with_testdb: Create a test database
    :return: None
    """
    db.drop_all()
    db.create_all()

    if with_testdb:
        db_uri = '{0}_test'.format(app.config['SQLALCHEMY_DATABASE_URI'])

        if not database_exists(db_uri):
            create_database(db_uri)

    return None


@click.command()
def seed():
    """
    Seed the database with an initial user.

    :return: User instance
    """
    if User.find_by_identity(app.config['SEED_ADMIN_EMAIL']) is not None:
        return None

    params = {
        'role': 'admin',
        'email': app.config['SEED_ADMIN_EMAIL'],
        'password': app.config['SEED_ADMIN_PASSWORD']
    }

    return User(**params).save()


def seed_meter():
    """
    Seed the database with an initial user.

    :return: User instance
    """
    if Meter.find_by_identity('CLE154179105777') is not None:
        return None

    params = {
        'serial_number': 'CLE154179105777',
        'phone_number': '00861064685239932',
        'sequence_number': 'china 19'
    }

    return Meter(**params).save()


def seed_alert():
    """
    Seed the database with an initial user.

    :return: User instance
    """
    meter = Meter.find_by_identity('CLE154179105777')
    if meter is not None:
        random_alerts = []
        data = []
        for i in range(0, 99):
            random_alerts.append(fake.slug())

        while True:
            if len(random_alerts) == 0:
                break
            fake_datetime = fake.date_time_between(
                start_date='-1y', end_date='now')

            created_on = fake_datetime

            fake_datetime = fake.date_time_between(
                start_date='-1y', end_date='now')

            capture_time = fake_datetime
            alert = random_alerts.pop()
            params = {
                'code': str(int(round((random.random() * 1000)))),
                'name': alert,
                'capture_time': capture_time,
                'meter_id': meter.id,
                'created_on': created_on
            }
            data.append(params)

        return _bulk_insert(AlertData, data, 'alerts')
    return None


def seed_imei():
    data = []
    wb = load_workbook(filename='data/IMEI number CL710K22.xlsx', read_only=True)
    click.echo(wb.sheetnames)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.rows:
            if row[0].value and row[0].value != "Meter equipment identifier":
                val = None
                if row[6].value:
                    val = str(row[6].value)

                params = {
                    'meter_identifier': "CLE1" + str(row[0].value) + "0",
                    'modem_imei': val
                }
                data.append(params)

    wb = load_workbook(filename='data/IMEI number CL730S22.xlsx', read_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.rows:
            if row[0].value and row[0].value != "Meter equipment identifier":
                val = None
                if row[6].value:
                    val = str(row[6].value)

                params = {
                    'meter_identifier': "CLE3" + str(row[0].value) + "0",
                    'modem_imei': val
                }
                data.append(params)
    return _bulk_insert(Imei, data, 'imei')


def seed_sim():
    data = []
    wb = load_workbook(filename='data/Beco Numbers.xlsx', read_only=True)
    click.echo(wb.sheetnames)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.rows:
            if row[1].value and row[1].value != "MSISDN":
                params = {
                    'phone_number': "00252"+str(row[1].value),
                    'sim_serial': str(row[2].value)
                }
                data.append(params)
    return _bulk_insert(Sim, data, 'sim')
    return None


@click.command()
@click.option('--with-testdb/--no-with-testdb', default=False,
              help='Create a test db too?')
@click.pass_context
def reset(ctx, with_testdb):
    """
    Init and seed automatically.

    :param ctx:
    :param with_testdb: Create a test database
    :return: None
    """
    ctx.invoke(init, with_testdb=with_testdb)
    ctx.invoke(seed)
    ctx.invoke(seed_meter)
    ctx.invoke(seed_alert)
    ctx.invoke(seed_imei)
    ctx.invoke(seed_sim)

    return None


def _bulk_insert(model, data, label):
    """
    Bulk insert data to a specific model and log it. This is much more
    efficient than adding 1 row at a time in a loop.

    :param model: Model being affected
    :type model: SQLAlchemy
    :param data: Data to be saved
    :type data: list
    :param label: Label for the output
    :type label: str
    :return: None
    """
    with app.app_context():
        model.query.delete()
        db.session.commit()
        db.engine.execute(model.__table__.insert(), data)
        _log_status(model.query.count(), label)

    return None


cli.add_command(init)
cli.add_command(seed)
cli.add_command(reset)
